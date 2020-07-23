# -*- coding: utf-8 -*-
#@Time      :2020/6/15    23:19
#@Author    :xj
#@Email     :1027867874@qq.com
#@File      :login_data.py
#@Software  :PyCharm

from openpyxl import load_workbook
from d2020_07_01.common.EnvData import EnvData


class ReadExcel(object):
    def __init__(self,file_path,sheet_name):
        """
        读取excel中的用例
        :param file_path: 加载目录下的xlsx文件
        :param sheet_name: sheet表单
        """
        self.file_path = file_path
        self.wb = load_workbook(self.file_path)
        self.sheet = self.wb[sheet_name]

    def load_titles(self):
        # 加载首行数据 作为字典的key
        case_key = list ()  # 收集excel表中首行数据作为key
        for item in list (self.sheet.rows)[0]:
            case_key.append (item.value)
        return case_key

    def load_data(self):
        '加载excel数据'
        case_key = self.load_titles()
        #第一种方法 （灵活，维护成本低）
        line_list = list() # 收集整个sheet表中的数据 每一个元素都是字典
        for cell in list(self.sheet.rows)[1:]: # 获取每行数据 类型为元祖
            line_dict = dict()
            for index in range(len(cell)):
                case_value= cell[index].value
                line_dict[case_key[index]] = case_value
            # 每条用例是实例化一个对象放在字典中为了进行对表格中参数化 做对象标记
            line_dict["obj"] = EnvData()
            line_list.append(line_dict)
        return line_list


        # 第二种方法 （麻烦，字段被写死）
        # line_list = list ()
        # for cell in list (self.sheet.rows)[1:]:
        #     cell_list = list()
        #     for item in cell:
        #         cell_list.append(item.value)
        #     line_dict = dict(zip(case_key,cell_list))
        #
        #     for key,value in line_dict.items():
        #         if isinstance(value,str):
        #             if value == "None": # 因为传参时要考虑None 所以需要进行特殊处理
        #                 line_dict[key] = None
        #             elif key == 'expect_result':
        #                 line_dict['expect_result'] = eval (line_dict['expect_result'])
        #             else:
        #                 line_dict[key] = line_dict[key]
        #         else:
        #             line_dict[key] = str(line_dict[key]) # 因为传参必须都是字符串 所以必须对数字类型的密码进行处理
        #
        #     line_list.append(line_dict)
        # return line_list




    def write_back_data(self,row,colum,input_value):
        '回写数据'
        self.sheet.cell(row,colum).value = input_value
        self.wb.save(self.file_path)

    def close_excel(self):
        self.wb.close ()
if __name__ == '__main__':
    # datas = ReadExcel (r"E:\py30\d2020_07_01\test_data\nmb_qcd.xlsx", "register").load_data ()
    # print(datas)
    # import json
    # param='{\n"mobile_phone": "13888889209"\n}'
    # print(json.loads(param))
    s = {'key1':'value1','key2':'value2'}
    print(s.keys())


